#!/usr/bin/env python3
"""
Transform VA-collected HOA Amenity Data from Excel to communities.json format.

This script reads the Excel file containing 59 communities with amenity data
collected by VA research team and transforms it into the communities.json format
used by the Loudoun Property Intelligence Platform.

Usage:
    python scripts/transform_va_amenity_data.py [--sample] [--merge]

    --sample: Show sample output for 3 communities (default)
    --merge:  Merge into communities.json (requires explicit flag)
"""

import json
import os
import re
import sys
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl


# File paths
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)
EXCEL_PATH = os.path.join(BASE_DIR, 'data/loudoun/communities/HOA_Amenity_Data_Collection (1).xlsx')
COMMUNITIES_JSON_PATH = os.path.join(BASE_DIR, 'data/loudoun/config/communities.json')
BACKUP_PATH = os.path.join(BASE_DIR, 'data/loudoun/config/communities.json.backup')


def parse_yes_no(value: Any) -> bool:
    """Convert Y/N/Yes/No values to boolean."""
    if value is None:
        return False
    val_str = str(value).strip().upper()
    return val_str in ('Y', 'YES', 'TRUE', '1')


def parse_count(value: Any) -> Optional[int]:
    """Parse count values, handling ~N and N+ notation."""
    if value is None or value == '':
        return None
    val_str = str(value).strip()
    # Remove ~ prefix (approximate)
    val_str = val_str.lstrip('~')
    # Remove + suffix
    val_str = val_str.rstrip('+')
    try:
        # Try to parse as float first (Excel sometimes stores as float)
        return int(float(val_str))
    except (ValueError, TypeError):
        return None


def parse_trail_miles(value: Any) -> Optional[float]:
    """Parse trail miles, handling ~N notation."""
    if value is None or value == '':
        return None
    val_str = str(value).strip()
    val_str = val_str.lstrip('~')
    try:
        return float(val_str)
    except (ValueError, TypeError):
        return None


def name_to_key(name: str) -> str:
    """Convert community display name to JSON key format.

    Examples:
        "Ashburn Farm" -> "ashburn_farm"
        "One Loudoun" -> "one_loudoun"
        "Courts and Ridges of Ashburn" -> "courts_and_ridges_of_ashburn"
    """
    # Convert to lowercase
    key = name.lower()
    # Replace spaces and special chars with underscores
    key = re.sub(r'[^a-z0-9]+', '_', key)
    # Remove leading/trailing underscores
    key = key.strip('_')
    # Collapse multiple underscores
    key = re.sub(r'_+', '_', key)
    return key


def transform_excel_row(row: tuple, headers: List[str]) -> Dict[str, Any]:
    """Transform a single Excel row into communities.json format."""
    # Create a dict from headers and row values
    data = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}

    community_name = data.get('Community Name', '')
    if not community_name:
        return None

    # Build the community structure
    community = {
        'display_name': community_name,
        'data_source': 'VA Research Team',
        'data_verified': data.get('Date Verified'),
        'verified_by': data.get('Verified By'),
        'amenities': {
            'pools': parse_count(data.get('Pool Count')) or (1 if parse_yes_no(data.get('Has Pool')) else 0),
            'has_pool': parse_yes_no(data.get('Has Pool')),
            'tennis_courts': parse_count(data.get('Tennis Courts')) or (1 if parse_yes_no(data.get('Has Tennis')) else 0),
            'has_tennis': parse_yes_no(data.get('Has Tennis')),
            'clubhouse': parse_yes_no(data.get('Has Clubhouse')),
            'fitness_center': parse_yes_no(data.get('Has Fitness Center')),
            'has_trails': parse_yes_no(data.get('Has Trails')),
            'trail_miles': parse_trail_miles(data.get('Trail Miles')),
            'playgrounds': parse_count(data.get('Playground Count')) or (1 if parse_yes_no(data.get('Has Playground')) else 0),
            'has_playground': parse_yes_no(data.get('Has Playground')),
            'basketball': parse_yes_no(data.get('Has Basketball')),
            'dog_park': parse_yes_no(data.get('Has Dog Park')),
            'pickleball': parse_yes_no(data.get('Has Pickleball')),
        }
    }

    # Add HOA website if available
    hoa_website = data.get('HOA Website')
    if hoa_website and str(hoa_website).strip():
        community['hoa_website'] = str(hoa_website).strip()

    # Add notes if available
    notes = data.get('Notes')
    if notes and str(notes).strip():
        community['notes'] = str(notes).strip()

    return community


def read_excel_data() -> List[Dict[str, Any]]:
    """Read and transform all communities from Excel file."""
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    # Get headers from first row
    headers = [cell.value for cell in ws[1]]

    communities = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:  # Community Name is in column 1 (0-indexed)
            community = transform_excel_row(row, headers)
            if community:
                communities.append(community)

    return communities


def load_existing_communities() -> Dict[str, Any]:
    """Load existing communities.json file."""
    with open(COMMUNITIES_JSON_PATH, 'r') as f:
        return json.load(f)


def merge_communities(existing: Dict[str, Any], new_communities: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Merge new communities into existing data structure.

    Rules:
    - New communities are added with their full data
    - Existing communities: update amenities only, preserve other data
    """
    result = existing.copy()

    # Get existing community display names for comparison
    existing_display_names = {}
    for key, data in result.get('communities', {}).items():
        display_name = data.get('display_name', key)
        existing_display_names[display_name] = key

    added = 0
    updated = 0

    for new_comm in new_communities:
        display_name = new_comm['display_name']
        key = name_to_key(display_name)

        if display_name in existing_display_names:
            # Update existing community's amenities
            existing_key = existing_display_names[display_name]
            existing_amenities = result['communities'][existing_key].get('amenities', {})

            # Merge amenities (new data takes precedence for VA-verified fields)
            for amenity_key, amenity_val in new_comm['amenities'].items():
                if amenity_val is not None and amenity_val != 0 and amenity_val != False:
                    existing_amenities[amenity_key] = amenity_val

            result['communities'][existing_key]['amenities'] = existing_amenities
            result['communities'][existing_key]['va_verified'] = True
            result['communities'][existing_key]['va_verified_date'] = new_comm.get('data_verified')

            # Add notes if not present
            if new_comm.get('notes') and not result['communities'][existing_key].get('notes'):
                result['communities'][existing_key]['notes'] = new_comm['notes']

            updated += 1
        else:
            # Add new community
            result['communities'][key] = new_comm
            added += 1

    # Update metadata
    result['_metadata']['last_updated'] = datetime.now().strftime('%Y-%m-%d')
    result['_metadata']['va_data_integrated'] = True
    result['_metadata']['va_communities_added'] = added
    result['_metadata']['va_communities_updated'] = updated

    return result


def show_sample_output(communities: List[Dict[str, Any]], count: int = 3):
    """Display sample transformed communities for review."""
    print(f"\n{'='*60}")
    print(f"SAMPLE OUTPUT: First {count} Transformed Communities")
    print(f"{'='*60}\n")

    for i, comm in enumerate(communities[:count], 1):
        key = name_to_key(comm['display_name'])
        print(f"--- Community {i}: {comm['display_name']} ---")
        print(f"Key: \"{key}\"")
        print(json.dumps({key: comm}, indent=2))
        print()


def main():
    """Main entry point."""
    args = sys.argv[1:]

    print("=" * 60)
    print("VA Amenity Data Transformation Script")
    print("=" * 60)

    # Read Excel data
    print(f"\nReading Excel file: {EXCEL_PATH}")
    communities = read_excel_data()
    print(f"Parsed {len(communities)} communities from Excel")

    # Show sample by default
    if '--merge' not in args:
        show_sample_output(communities)
        print("=" * 60)
        print("To merge into communities.json, run with --merge flag")
        print("=" * 60)
        return

    # Merge mode
    print("\nLoading existing communities.json...")
    existing = load_existing_communities()
    print(f"Existing communities: {len(existing.get('communities', {}))}")

    print("\nMerging data...")
    merged = merge_communities(existing, communities)
    print(f"Total communities after merge: {len(merged.get('communities', {}))}")

    # Write merged data
    print(f"\nWriting to: {COMMUNITIES_JSON_PATH}")
    with open(COMMUNITIES_JSON_PATH, 'w') as f:
        json.dump(merged, f, indent=2)

    print("\n" + "=" * 60)
    print("MERGE COMPLETE")
    print(f"Added: {merged['_metadata'].get('va_communities_added', 0)} new communities")
    print(f"Updated: {merged['_metadata'].get('va_communities_updated', 0)} existing communities")
    print("=" * 60)


if __name__ == '__main__':
    main()
